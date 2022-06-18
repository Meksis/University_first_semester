# Импорты библиотек
import openpyxl		# pip install openpyxl , библа для работы с таблицей
from openpyxl import *		# вытаскиваем из нее все вложенные методы
import folium	# pip install folium , библа для работы с картой
from folium import *	# вытаскиваем из нее все вложенные методы
import io 	# Библиотека для отрисовки карты в проге
import sys
from PyQt5.QtWidgets import *	# pip install pyqt5 , библиотека для создания интерфейса
from PyQt5.QtWebEngineWidgets import * 		# pip install PyQtWebEngine
from PyQt5.QtCore import *


#xl_path=input('Введите полный путь до файла (вместе с расширением файла) ')		# Можно заменить на статический путь, если не хотите вручную вводить
"""xl_path_=''
for sym in xl_path:
	if sym != '\\':
		xl_path_+=sym
	else:
		xl_path_+='/'"""

xl_path='J:/Downloads/Telegram Desktop/Транзит 2019-2020 гг..xlsx'
#geojson_path='J:/Downloads/Загрузки Opera GX/map (2).geojson'

xl_file = load_workbook(filename=xl_path,  data_only=True)	# Создание мелкой копии файла. Атрибут data_only позволяет избежать получения формул при попытке получить данные из ячейки
working_sheet=xl_file[xl_file.sheetnames[0]]	# Выбор таблицы. xl_file.sheetnames возвращает список с листами таблицы. Указывая [0] мы передаем проге инфу название нужного листа и он полнстью записывается в working_sheet

countries=[]

def map_update(x, y, popup, map_, color, zoom=3):	# Функция для добавления маркеров по указанным координатам на карту, так же указанную при вызове функции
	return(folium.Marker(location=[x, y], icon=folium.Icon(color = color.lower()), popup=popup).add_to(map_))	
	#return(Map(location=[x, y], zoom_start=zoom))


def msg_showing(main_text, secondary_text):		# Функция для вызова информационного окна с указанными заголовком и основным текстом
	msg = QMessageBox()
	msg.setWindowTitle(main_text)
	msg.setText(secondary_text)
	#msg.setIcon(QMessageBox.Warning)
	msg.exec_()

def search_button():	# Функция, описывающая действия при нажатии кнопки "Запрос" в окне поиска. При создании интерфейса можно будет удалить к чертям
	text=line_edit.text().upper()	# Получаем текст с поляя вода текста, переводим в верхний регистр ( в таблице все страны написаны капсом )
	counter=0
	parts_amount=0 		# переменна для хранения количества слов, разделенных пробелом
	country=''
	countries=[]
	for symbol in text:
		if symbol!=' ':		# Поскольку данные разделены пробелом, по нему и будем определять конец слова
			country+=symbol # Поскольку символ НЕ пробел, дописываем его в переменную country
			counter+=1 		# Счетчик пригодится при записывании данных после пробела
		else:
			countries.append(country)
			countries.append(text[counter+1:len(text)])		# Записываем в список все символы, следующие после пробела. Да, это костыль, но от этого избавимся при написании хоть какого интерфейса
			parts_amount+=1
			break


	if parts_amount!=0:		# Если в поле ввода было введено два элемента или больше 
		country_out=countries[0]								# Страны, полученные при парсинге данных из онка ввода текста
		country_in=countries[1]

		country_out_f=''										# Переменные для хранения стран, которые перебирает цикл ниже
		country_in_f=''
		switcher=0 												# Когда находится нужное значение, эта переменная примет значение 1, что позволит отлавливать случаи, когда данные не найдены
		find_counter=2 											# Переменная для хранения номера строки с найденными совпадениями
		data_add_out=''

		for row_number in range(find_counter, working_sheet.max_row+1):	# Перебираем номера всех строк таблицы. Начинаем со второй строки потому, что в openpyxl нумерация строк и столбцов начинается с 1, а первая строка - название столбца

			country_out_f=str(working_sheet.cell(row=row_number, column=3).value)	# Присваиваем значения ячеек с номером строки row_number и номерами столбов 3 и 8 переменным для поиска совпадений. В 3 и 8 стобцах находятся страны отправления и прибытия соответственно.
			country_in_f=str(working_sheet.cell(row=row_number, column=8).value)

			if country_in==country_in_f and country_out==country_out_f:				# если страна прибытия == найденной стране прибытия И если страна отправления == найденной стране отправления
				data_add=[]
				#data_add_out=''
				for column_number in range(1, 17):										# Перебираем номера столбцов с полезной информацией
					data_add.append([working_sheet.cell(row=1, column=column_number).value, working_sheet.cell(row=row_number, column=column_number).value])	# Дописываем в список название столбца и информацию из ячейки

				for data_list in data_add:	# Готовим данные для вывода в окно
					data_add_out+=f'{data_list[0]}: {data_list[1]}\n'	# Собсна заголовок и информация. \n в конце позволяет переносить последующий текст на строку ниже
				data_add_out+='------------------------\n'
				switcher=1
				find_counter=row_number
				#break	# Выходим из цикла. Если этого не сделать, то даже при нахождении нужной комбинации значений, мы будем сравнивать все последующие данные в таблице

		if not switcher:											# Если не удалось найти данных по запросу
			msg_showing('Результаты поиска', 'Совпадений не найдено')	# Выводим окно с уведомлением о неудачном поиске инфы
		else:
			print(row_number, row_number)
			msg_showing('Результаты поиска', data_add_out)	# Выхываем окно с найденной информацией
	else:														# Если в поле ввода была пустота или меньше двух элементов
		msg_showing('Ошибка', 'Введено недостаточно стран')		



for row_number in range(3, 241+1):	# Проходимся по столбцам справа таблицы
	countries.append([working_sheet.cell(row=row_number, column=28).value, working_sheet.cell(row=row_number, column=29).value, working_sheet.cell(row=row_number, column=30).value])
	# Вносим в список стран название страны, ее регион и координаты. Если чего-то нет, то будет внесено Null

countries.sort()	# Сортируем страны в алфавитном порядке

coordinates=[]
mid_list=[]

for list_ in countries:		# Перебираем списки из массива, созданного выше

	if list_[2]!=None:		# Если графа координат не Null, то есть, если в таблице были указаны координаты
		number=''				# Переменная для хранения символов-чисел координат
		for counter, symbol in enumerate(list_[2]):		# Координаты из тыблицы подтягиваются как строки вида XX.XXXXXXXX, XX.XXXXXXXX, а нам нужны float-числа. Ниже парсим данные для получения нужного результата. enumerate позволяет при каждом проходе цикла увеличивать на 1 или нужное число переменную, первой указанную после слова for.
			if symbol==',':						# Если символ оказался запятой (разделителем интересующих нас данных)
				mid_list.append(float(number))		# Дописываем в промежуточный список строку с координатой, конвернитрованную в float-число
				number=''							# Обнуляем переменную, хранящую строковое число-координату
			elif symbol==' ':					# Если символ - пробел, то пропускаем действие и ничего не делаем (pass позволяет пропустить действие и продолжить цикл без изменений)
				pass
			else:								# Если символ не пробел и не запятая, то бишь - число
				if counter==len(list_[2])-1:		# Если счётчик дошел до последнего символа, то есть, до последнего числа в строке
					mid_list.append(float(number))		# Дописываем в наш список промежуточных значений вторую координату

				else:								# Иначе
					number+=symbol						# Дописываем символ к уже имеющимся
		mid_list.append(list_[0])			# Так же дописываем в список с координатами название страны, что пригодится при расставлении маркеров
		coordinates.append(mid_list)		# Добавляем к списку координат получившийся список
		mid_list=[]							# Обнуляем список



map1=Map(location=[0, 0], zoom_start=3)		# Создаем объект карты (поскольку мы прописали from folium import *, то есть, импортировали все содержимое главного модуля folium, а метод Map относится именно к нему, то мы можем не прописывать folium.Map, а писать сразу Map)
for country_data in coordinates:			# Поочередно выбираем списки с информацией для построения маркеров
	map_update(country_data[0], country_data[1], country_data[2], map1, 'red')		# Вызываем функцию доя нанесения маркера на карту, передавая ей координаты (2), название страны, объект карты и цвет маркера
map1.save('map.html')	# Сохранение карты в директории, где находится питоновский файл


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setWindowTitle('Data Mine (DM)')
        #self.map_draw()

    def map_draw(self):
    	map_layout = QVBoxLayout(self)
    	engine = QWebEngineView(self)
    	data = io.BytesIO()				# Не спрашивайте, так было на стаке
    	map1.save(data, close_file=False)	# Тут тоже

    	map_layout = QVBoxLayout(self)

    	engine = QWebEngineView(self)
    
    	engine.setHtml(data.getvalue().decode())

    	#self.result_widget = SearchWidget()
    	#self.result_widget.show() 
        



if __name__ == "__main__":		# Необязательная конструкция. Позволяет проверять, открывается ли этот файл отдельно, пользователем, или программно

    app = QApplication(sys.argv)	# Создание объекта приложения

    

    main_window = MainWindow()

   
    #engine.load()


    map_layout.addWidget(engine)

    main_window.setLayout(map_layout)

    main_window.show()




    line_edit = QLineEdit()			# Объект поля ввода текста
    line_edit.setToolTip('Искомые страны через пробел ( 2 )')		# Подсказка при наведении на поле для ввода

    button = QPushButton('Запрос')		# Объект кнопки
    button.clicked.connect(search_button)		# Говорим проге, что надо вызывать функцию search_button при нажатии на кнопку
    #button.setToolTip('Simple button...')		# Подсказка при наведении на кнопку


    layout = QFormLayout()				# Обхект, который может сам позиционировать добавенные в него объекты в окне
    layout.addRow('Поиск:', line_edit)	# Добавляем поле ввода текста , делаем подсказку для юзера
    layout.addRow(button)				# Добавляем в окно кнопку
    #layout.addRow('Text edit:', text_edit)

    window = QWidget()					# Собственно, создаем само окно, в котором будет все наше поисковое богатство
    

    """window.setWindowTitle('Поиск')		# Устанавливаем название окна
                window.setLayout(layout)			# Добавляем объект с кнопкой и текстовым полем в окошко
                window.show()						# Показываем окошко. Можно будет показывать его при нажатии на кнопку в окне с картой"""



    #window.show()						# Отрисовываем окно поиска
    sys.exit(app.exec_())				# Завершаем программу при закрытии окон


    """
    data = io.BytesIO()				# Не спрашивайте, так было на стаке
    map1.save(data, close_file=False)	# Тут тоже

    map_shower = QWebEngineView()			# Создание объекта для просмотра html-файла нашей карты
    map_shower.setHtml(data.getvalue().decode())
    map_shower.resize(1920, 1080)			# Ресайзим окно по ШиринеХВысоте монитора.
    #w.show()						# Отрисовывем карту

    main_window.addWidget()
	



		area = QScrollArea(self)
        area.setWidgetResizable(True)
        self.scrollAreaWidgetContents = QLabel(text, self) 

        area.setWidget(self.scrollAreaWidgetContents)
        #button = QPushButton("Закрыть окно")
        #button.clicked.connect(self.goMainWindow) 

        layoutV = QVBoxLayout() 
        layoutV.addWidget(area)
        #layoutV.addWidget(button)
        self.setLayout(layoutV)
        """