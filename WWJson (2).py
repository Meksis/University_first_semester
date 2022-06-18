# Импорты библиотек
import openpyxl		# pip install openpyxl , библа для работы с таблицей
from openpyxl import *		# вытаскиваем из нее все вложенные методы
import folium	# pip install folium , библа для работы с картой
from folium import *	# вытаскиваем из нее все вложенные методы
import io 	# Библиотека для отрисовки карты в проге
import sys
from PyQt5 import QtGui
from PyQt5.QtWidgets import *	# pip install pyqt5 , библиотека для создания интерфейса
from PyQt5.QtWebEngineWidgets import * 		# pip install PyQtWebEngine
from PyQt5.QtCore import *

vadim_s_build=True		# !!!!!!!! ЕСЛИ У ВАС В ПРОГЕ ЭТА ПЕРЕМЕННАЯ True, ПОМЕНЯЙТЕ ЕЕ НА False ИЛИ ЗАМЕНИТЕ ПУТЬ ПРИ ВЫПОЛНЕНИИ УСЛОВИЯ НА СВОЙ !!!!!!!!

if vadim_s_build:
	xl_path='J:/Downloads/Telegram Desktop/Транзит 2019-2020 гг..xlsx'
else:
	xl_path=input('Введите полный путь до Excel-таблицы (вместе с расширением) ')


#geojson_path='J:/Downloads/Загрузки Opera GX/map (2).geojson'

xl_file = load_workbook(filename=xl_path,  data_only=True)	# Создание мелкой копии файла. Атрибут data_only позволяет избежать получения формул при попытке получить данные из ячейки
working_sheet=xl_file[xl_file.sheetnames[0]]	# Выбор таблицы. xl_file.sheetnames возвращает список с листами таблицы. Указывая [0] мы передаем проге инфу название нужного листа и он полнстью записывается в working_sheet
max_rows=working_sheet.max_row

countries=[]

def map_update(x, y, popup, map_, color, zoom=3):	# Функция для добавления маркеров по указанным координатам на карту, так же указанную при вызове функции
	return(folium.Marker(location=[x, y], icon=folium.Icon(color = color.lower()), popup=popup).add_to(map_))	
	#return(Map(location=[x, y], zoom_start=zoom))

def msg_showing(main_text, secondary_text):		# Функция для вызова информационного окна с указанными заголовком и основным текстом
	msg = QMessageBox()
	msg.setWindowTitle(main_text)
	msg.setText(secondary_text)
	#msg.setIcon(QMessageBox.Warning)
	msg.exec_()

def send_sheet_table(name, max_row_count=max_rows):
	for col_num in range(1, 16):
		if working_sheet.cell(row=1, column=col_num).value==name:
			break
	for eternal_number in range(2, max_row_count):
		print(working_sheet.cell(row=eternal_number, column=col_num).value)

def get_button_text(name):
	print(f'\n\"{name}\" button clicked\n')
	send_sheet_table(name, 100)

def window_construct(window, layout, values, rows_count=2):			# Для трех строчек надо сильно думать, ограничимся двумя
	if rows_count >= 3:
		msg_showing('Ошибка', 'Программа не поддерживает разбиение кнопок на три и более строки.')					# Нюансек
		sys.exit(0)

	else:

		grid_layout = layout

		move_row = len(values) - len(values) // rows_count			# + len(values) % rows_count		# 10 - 10 // 2 + 10 % 2 = 10 - 5 + 0 = 5             11 - 11 // 2 = 5		13 - 13 // 2 + 13 % 2 = 13 - 6 + 1
		row_counter=1

		for column_counter, number in enumerate(values):

			number = str(number)

			if move_row==column_counter+1:
				button=QPushButton(number)
				#button.resize(sizeHint(button))
				button.clicked.connect(lambda ch, name=number: get_button_text(name))
				grid_layout.addWidget(button, row_counter, column_counter)

				number = column_counter
				row_counter+=1

				for counter in range(0, column_counter):
					button=QPushButton(str(values[number+counter+1]))
					#button.resize(sizeHint(button))
					button.clicked.connect(lambda ch, name=str(values[number+counter+1]): get_button_text(name))
					grid_layout.addWidget(button, row_counter, counter)

				break

			else:
				#grid_layout.addWidget(QPushButton(number), row_counter, column_counter) 
				button=QPushButton(number)
				#button.resize(sizeHint(button))

				button.clicked.connect(lambda ch, name=number: get_button_text(name))


				grid_layout.addWidget(button, row_counter, column_counter)   
	return(move_row)
		
		#comboBox = QComboBox(window)
		#comboBox.addItem('Образец списка')
		#comboBox.move(20, 200)'''
		#grid_layout.addWidget(comboBox)


"""def search_button():	# Функция, описывающая действия при нажатии кнопки "Запрос" в окне поиска. При создании интерфейса можно будет удалить к чертям
	text=line_edit.text().upper()	# Получаем текст с поляя вода текста, переводим в верхний регистр ( в таблице все страны написаны капсом )
	counter=0
	parts_amount=0 		# переменна для хранения количества слов, разделенных пробелом
	country=''
	countries=[]
	for symbol in text:
		if symbol!=' ':		# Поскольку данные разделены пробелом, по нему и будем определять конец слова
			country+=symbol # Поскольку символ НЕ пробел, дописываем его в переменную country
			counter+=1 		# Счетчик пригодится при записывании данных после пробела
		else:
			countries.append(country)
			countries.append(text[counter+1:len(text)])		# Записываем в список все символы, следующие после пробела. Да, это костыль, но от этого избавимся при написании хоть какого интерфейса
			parts_amount+=1
			break


	if parts_amount!=0:		# Если в поле ввода было введено два элемента или больше 
		country_out=countries[0]								# Страны, полученные при парсинге данных из онка ввода текста
		country_in=countries[1]

		country_out_f=''										# Переменные для хранения стран, которые перебирает цикл ниже
		country_in_f=''
		switcher=0 												# Когда находится нужное значение, эта переменная примет значение 1, что позволит отлавливать случаи, когда данные не найдены
		find_counter=2
		data_add_out=''


		for row_number in range(find_counter, working_sheet.max_row+1):	# Перебираем номера всех строк таблицы. Начинаем со второй строки потому, что в openpyxl нумерация строк и столбцов начинается с 1, а первая строка - название столбца

			country_out_f=str(working_sheet.cell(row=row_number, column=3).value)	# Присваиваем значения ячеек с номером строки row_number и номерами столбов 3 и 8 переменным для поиска совпадений. В 3 и 8 стобцах находятся страны отправления и прибытия соответственно.
			country_in_f=str(working_sheet.cell(row=row_number, column=8).value)

			if country_in==country_in_f and country_out==country_out_f:				# если страна прибытия == найденной стране прибытия И если страна отправления == найденной стране отправления
				data_add=[]
				for column_number in range(1, 17):										# Перебираем номера столбцов с полезной информацией
					data_add.append([working_sheet.cell(row=1, column=column_number).value, working_sheet.cell(row=row_number, column=column_number).value])	# Дописываем в список название столбца и информацию из ячейки

				for data_list in data_add:	# Готовим данные для вывода в окно
					data_add_out+=f'{data_list[0]}: {data_list[1]}\n'	# Собсна заголовок и информация. \n в конце позволяет переносить последующий текст на строку ниже
				data_add_out+='------------------------\n'
				
				switcher=1
				find_counter=row_number
				#break	# Выхожим из цикла. Если этого не сделать, то даже при нахождении нужной комбинации значений, мы будем сравнивать все последующие данны в таблице
		if not switcher:											# Если не удалось найти данных по запросу
			msg_showing('Результаты поиска', 'Совпадений не найдено')	# Выводим окно с уведомлением о неудачном поиске инфы
		else:
			print('Типа окно, да')
			#result_window(data_add_out)	# Выхываем окно с найденной информацией
			#ResultWindow.show()
	else:														# Если в поле ввода была пустота или меньше двух элементов
		msg_showing('Ошибка', 'Введено недостаточно стран')		
"""


for row_number in range(3, 241+1):	# Проходимся по столбцам справа таблицы
	countries.append([working_sheet.cell(row=row_number, column=28).value, working_sheet.cell(row=row_number, column=29).value, working_sheet.cell(row=row_number, column=30).value])
	# Вносим в список стран название страны, ее регион и координаты. Если чего-то нет, то будет внесено Null




countries.sort()	# Сортируем страны в алфавитном порядке

coordinates=[]
mid_list=[]

for list_ in countries:		# Перебираем списки из массива, созданного выше

	if list_[2]!=None:		# Если графа координат не Null, то есть, если в таблице были указаны координаты
		number=''				# Переменная для хранения символов-чисел координат
		for counter, symbol in enumerate(list_[2]):		# Координаты из тыблицы подтягиваются как строки вида XX.XXXXXXXX, XX.XXXXXXXX, а нам нужны float-числа. Ниже парсим данные для получения нужного результата. enumerate позволяет при каждом проходе цикла увеличивать на 1 или нужное число переменную, первой указанную после слова for.
			if symbol==',':						# Если символ оказался запятой (разделителем интересующих нас данных)
				mid_list.append(float(number))		# Дописываем в промежуточный список строку с координатой, конвернитрованную в float-число
				number=''							# Обнуляем переменную, хранящую строковое число-координату
			elif symbol==' ':					# Если символ - пробел, то пропускаем действие и ничего не делаем (pass позволяет пропустить действие и продолжить цикл без изменений)
				pass
			else:								# Если символ не пробел и не запятая, то бишь - число
				if counter==len(list_[2])-1:		# Если счётчик дошел до последнего символа, то есть, до последнего числа в строке
					mid_list.append(float(number))		# Дописываем в наш список промежуточных значений вторую координату

				else:								# Иначе
					number+=symbol						# Дописываем символ к уже имеющимся
		mid_list.append(list_[0])			# Так же дописываем в список с координатами название страны, что пригодится при расставлении маркеров
		coordinates.append(mid_list)		# Добавляем к списку координат получившийся список
		mid_list=[]							# Обнуляем список

print(coordinates)

map1=Map(location=[0, 0], zoom_start=3)		# Создаем объект карты (поскольку мы прописали from folium import *, то есть, импортировали все содержимое главного модуля folium, а метод Map относится именно к нему, то мы можем не прописывать folium.Map, а писать сразу Map)

for country_data in coordinates:			# Поочередно выбираем списки с информацией для построения маркеров
	map_update(country_data[0], country_data[1], country_data[2], map1, 'red')		# Вызываем функцию доя нанесения маркера на карту, передавая ей координаты (2), название страны, объект карты и цвет маркера
map1.save('map.html')	# Сохранение карты в директории, где находится питоновский файл

main_columns=[]
for column_number in range(1, 16):
	main_columns.append(working_sheet.cell(row=1, column=column_number).value)


_translate=QCoreApplication.translate


class ResultWindow(QWidget):
	def __init__(self):								# "Магический" метод, позволяющий выполнять указанные действия при создании экземпляров класса
		super().__init__()
		self.initUI()								# При создании экземпляра вызываем функцию отрисовки интерфейса

	def search_button_reaction(self, font):			# Функция, описывающая действия при нажатии кнопки "Запрос" в окне поиска. При создании интерфейса надо будет подредачить
		text=self.line_edit.text().upper()	# Получаем текст с поляя вода текста, переводим в верхний регистр ( в таблице все страны написаны капсом )
		counter=0
		parts_amount=0 		# переменна для хранения количества слов, разделенных пробелом
		country=''
		countries=[]
		found_counter=0

		for symbol in text:
			if symbol!=' ':		# Поскольку данные разделены пробелом, по нему и будем определять конец слова
				country+=symbol # Поскольку символ НЕ пробел, дописываем его в переменную country
				counter+=1 		# Счетчик пригодится при записывании данных после пробела
			else:
				countries.append(country)
				countries.append(text[counter+1:len(text)])		# Записываем в список все символы, следующие после пробела. Да, это костыль, но от этого избавимся при написании хоть какого интерфейса
				parts_amount+=1
				break


		if parts_amount!=0:		# Если в поле ввода было введено два элемента или больше 
			country_out=countries[0]								# Страны, полученные при парсинге данных из онка ввода текста
			country_in=countries[1]

			country_out_f=''										# Переменные для хранения стран, которые перебирает цикл ниже
			country_in_f=''
			switcher=0 												# Когда находится нужное значение, эта переменная примет значение 1, что позволит отлавливать случаи, когда данные не найдены
			find_counter=2 											# Переменная, позволяющая реализовывать смещение в цикле поиска совпадений для нахождения всех совпадений. Без этого будет выведено последнее совпадение, а не все
			data_add_out=''											# Переменная для хранения отформатированных найденных совпадений


			for row_number in range(find_counter, working_sheet.max_row+1):	# Перебираем номера всех строк таблицы. Начинаем со второй строки потому, что в openpyxl нумерация строк и столбцов начинается с 1, а первая строка - название столбца

				country_out_f=str(working_sheet.cell(row=row_number, column=3).value)	# Присваиваем значения ячеек с номером строки row_number и номерами столбов 3 и 8 переменным для поиска совпадений. В 3 и 8 стобцах находятся страны отправления и прибытия соответственно.
				country_in_f=str(working_sheet.cell(row=row_number, column=8).value)

				if country_in==country_in_f and country_out==country_out_f:				# если страна прибытия == найденной стране прибытия И если страна отправления == найденной стране отправления
					data_add=[]															# Список для хранения пар Имя столбца - значение в форме, удобной для форматирования ниже
					found_counter+=1 													# Увеличиваем счетчик найденных совпадений
					for column_number in range(1, 17):										# Перебираем номера столбцов с полезной информацией
						if column_number!=11:
							data_add.append([working_sheet.cell(row=1, column=column_number).value, working_sheet.cell(row=row_number, column=column_number).value])	# Дописываем в список название столбца и информацию из ячейки
						else:
							data_add.append([working_sheet.cell(row=1, column=column_number).value, round(working_sheet.cell(row=row_number, column=column_number).value, 2)])	

					data_add_out+=f'\n\n-------- {found_counter} результат --------\n'
					for data_list in data_add:	# Готовим данные для вывода в окно
						data_add_out+=f'\n{data_list[0]}: {data_list[1]}'	# Собсна заголовок и информация. \n в конце позволяет переносить последующий текст на строку ниже
					data_add_out+=f'\n\n---------------------------------------------\n'
				
					switcher=1 												# Говорим проге, что у нас есть хотя бы одно совпадение
					find_counter=row_number									# Сдвигаем начало цикла for на найденный нами номер строки
					#break	# Выхожим из цикла. Если этого не сделать, то даже при нахождении нужной комбинации значений, мы будем сравнивать все последующие данны в таблице
			if not switcher:											# Если не удалось найти данных по запросу
				msg_showing('Результаты поиска', 'Совпадений не найдено')	# Выводим окно с уведомлением о неудачном поиске инфы
			else:
				#self.text_label.setText(data_add_out)
				self.scrollAreaWidgetContents = QLabel(f'Найдено совпадений: {found_counter}'+data_add_out, self) 	# Изменяем текст прокручиваемого окна на найденные страны
				self.area.setWidget(self.scrollAreaWidgetContents)														# Реализуем внесение текста
		else:														# Если в поле ввода была пустота или меньше двух элементов
			msg_showing('Ошибка', 'Введено недостаточно стран')	


	def initUI(self):								# Функция для отрисовки интерфейса
		self.resize(int(screen_w/2), int(screen_h/2))	# Ресайз окна под половину ширины и высоты монитора
		font = QtGui.QFont()							# Указываем шрифт текста
		font.setPointSize(14)

		self.line_edit = QLineEdit(self)			# Объект поля ввода текста. Указываем, что он принадлежит к объекту QWidget класса ResultWindow, оно же - self
		self.line_edit.setFont(font)				# Форматируем объект. В данном случае - только меняем размер шрифта
		self.line_edit.setToolTip('Искомые страны через пробел ( 2 )')		# Подсказка при наведении на поле для ввода

		self.button = QPushButton('Запрос', self)		# Объект кнопки. Указываем текст кнопки и ее принадлежность к виджету
		self.button.setFont(font)					# Форматируем объект. В данном случае - только меняем размер шрифта
		self.button.clicked.connect(self.search_button_reaction)		# Говорим проге, что надо вызывать функцию search_button при нажатии на кнопку
		
		self.area = QScrollArea(self)				# Создание объекта, способного реализовывать прокрутку своего содержимого. При множестве найденных результатов поиска это - лучшее решение
		self.area.setFont(font)						# Форматируем объект. В данном случае - только меняем размер шрифта
		self.area.setWidgetResizable(True)			# Говорим проге, что содержимое можно прокручивать

		self.layout = QFormLayout(self)	# Объект, который может сам позиционировать добавенные в него объекты в окне
		self.layout.addRow('Поиск:', self.line_edit)	# Добавляем поле ввода текста , делаем подсказку для юзера
		self.layout.addRow(self.button)					# Добавляем кнопку
		#self.layout.addRow(self.text_label)
		self.layout.addRow(self.area)					# Добавляем поле прокрутки

class MapWindow(QWidget):
	def __init__(self, buffer_data, column_names):								# "Магический" метод, позволяющий выполнять указанные действия при создании экземпляров класса
		super().__init__()
		self.buffer_data = buffer_data
		self.screen_w = screen_w
		self.screen_h = screen_h
		self.column_names = column_names
		self.initUI()								# При создании экземпляра вызываем функцию отрисовки интерфейса

	def initUI(self):								# Функция для отрисовки интерфейса
		#self.resize(int(screen_w/2), int(screen_h/2))	# Ресайз окна под половину ширины и высоты монитора
		self.resize(screen_w, screen_h)			# Ресайзим окно по ШиринеХВысоте монитора. 
		font = QtGui.QFont()							# Указываем шрифт текста
		font.setPointSize(14)

		self.grid_layout = QGridLayout(self)

		self.map_view = QWebEngineView(self)			# Создание объекта для просмотра html-файла нашей карты
		self.map_view.setHtml(self.buffer_data.getvalue().decode())
		self.map_view.setWindowTitle('Карта Мира')

		self.move = window_construct(self, self.grid_layout, self.column_names)

		self.grid_layout.addWidget(self.map_view, 5, 0, 10, self.move)


if __name__ == "__main__":		# Необязательная конструкция. Позволяет проверять, открывается ли этот файл отдельно, пользователем, или программно

    app = QApplication(sys.argv)							# Создание объекта приложения
    screen = app.primaryScreen().availableGeometry()		# Получаем значения доступного для использования пространства монитора 
    screen_w=screen.width()									# Записываем доступную ширину монитора
    screen_h=screen.height()								# Записываем доступную высоту монитора

    data = io.BytesIO()				# Не спрашивайте, так было на стаке
    map1.save(data, close_file=False)	# Тут тоже

    map_window = MapWindow(data, main_columns)
    map_window.setObjectName('MapWindow')
    map_window.setWindowTitle('Карта Мира')
    map_window.show()

    result_window = ResultWindow()							# Создаем экземпляр класса ResultWindow, отвечающего за выведение окна для поиска и выведения результатов поиска
    result_window.setObjectName('ResultWindow')				# Присваиваем экземпляру внутреннее программное имя
    result_window.setWindowTitle('Поиск')					# Меняем заголовок окна
    result_window.show()									# Выводим диалоговое окно

    """line_edit = QLineEdit()			# Объект поля ввода текста
                line_edit.setToolTip('Искомые страны через пробел ( 2 )')		# Подсказка при наведении на поле для ввода
            
                button = QPushButton('Запрос')		# Объект кнопки
                button.clicked.connect(search_button)		# Говорим проге, что надо вызывать функцию search_button при нажатии на кнопку
                #button.setToolTip('Simple button...')		# Подсказка при наведении на кнопку
            
            
                layout = QFormLayout()				# Обхект, который может сам позиционировать добавенные в него объекты в окне
                layout.addRow('Поиск:', line_edit)	# Добавляем поле ввода текста , делаем подсказку для юзера
                layout.addRow(button)				# Добавляем в окно кнопку
                #layout.addRow('Text edit:', text_edit)
            
                window = QWidget()					# Собственно, создаем само окно, в котором все наше поисковое богатство будет
                window.setWindowTitle('Поиск')		# Устанавливаем название окна
                window.setLayout(layout)			# Добавляем объект с кнопкой и текстовым полем в окошко
                #window.show()						# Показываем окошко. Можно будет показывать его при нажатии на кнопку в окне с картой
            
            
            
            
                #window.show()						# Отрисовываем окно поиска"""

    sys.exit(app.exec_())				# Завершаем программу при закрытии окон