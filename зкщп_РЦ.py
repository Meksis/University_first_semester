import openpyxl     # pip install openpyxl , библа для работы с таблицей
from openpyxl import *      # вытаскиваем из нее все вложенные методы
import folium   # pip install folium , библа для работы с картой
from folium import *    # вытаскиваем из нее все вложенные методы
import io   # Библиотека для отрисовки карты в проге
import os
from os import *
import sys
from PyQt5 import QtGui
from PyQt5.QtWidgets import *   # pip install pyqt5 , библиотека для создания интерфейса
from PyQt5.QtWebEngineWidgets import *      # pip install PyQtWebEngine
from PyQt5.QtCore import *

values_dict={}
countries=[]
coordinates=[]
mid_list=[]

_translate=QCoreApplication.translate


vadim_s_build=True      # !!!!!!!! ЕСЛИ У ВАС В ПРОГЕ ЭТА ПЕРЕМЕННАЯ True, ПОМЕНЯЙТЕ ЕЕ НА False ИЛИ ЗАМЕНИТЕ ПУТЬ ПРИ ВЫПОЛНЕНИИ УСЛОВИЯ НА СВОЙ !!!!!!!!

if vadim_s_build:
    xl_path='J:/Downloads/Telegram Desktop/Транзит 2019-2020 гг..xlsx'
else:
    xl_path=input('Введите полный путь до Excel-таблицы (вместе с расширением) ')

xl_file = load_workbook(filename=xl_path,  data_only=True)  # Создание мелкой копии файла. Атрибут data_only позволяет избежать получения формул при попытке получить данные из ячейки
working_sheet=xl_file[xl_file.sheetnames[0]]    # Выбор таблицы. xl_file.sheetnames возвращает список с листами таблицы. Указывая [0] мы передаем проге инфу название нужного листа и он полнстью записывается в working_sheet
max_rows=working_sheet.max_row

def msg_showing(main_text, secondary_text):     # Функция для вызова информационного окна с указанными заголовком и основным текстом
    msg = QMessageBox()
    msg.setWindowTitle(main_text)
    msg.setText(secondary_text)
    #msg.setIcon(QMessageBox.Warning)
    msg.exec_()

def sizeHint(widget):
    a = widget.sizeHint()
    a.setHeight(250)
    a.setWidth(50)
    return a

def map_update(x, y, popup, map_, color, zoom=3):   # Функция для добавления маркеров по указанным координатам на карту, так же указанную при вызове функции
    return(folium.Marker(location=[x, y], icon=folium.Icon(color = color.lower()), popup=popup).add_to(map_))   
    #return(Map(location=[x, y], zoom_start=zoom))

def column_dicts():

    for column_number in range(1, working_sheet.max_column+1):
        if working_sheet.cell(row=1, column=column_number).value == None:
            break
        else:
            column_values=[]
            for row_number in range(2, working_sheet.max_row+1):
                value=working_sheet.cell(row=row_number, column=column_number).value

                if value == None:
                    break

                else:
                    if value not in column_values:
                        if isinstance(value, float):
                            value=round(value, 2)
                        column_values.append(value)
            column_values.sort()
            values_dict.update({working_sheet.cell(row=1, column=column_number).value : column_values})

class mainWindow(QWidget):
    def __init__(self, values, coordinates, map_, map_window): # Названия столбоцв
        super().__init__()
        self.rows_count=2
        self.values = values 
        self.coordinates = coordinates
        self.map_ = map_
        self.map_window = map_window
        self.resize(int(screen_w/2), int(screen_h/2))  # Ресайз окна под половину ширины и высоты монитора
        #self.resize(int(1920/2), int(1080/2))

        self.window_construct()

    def index_change_reaction(self, object_name, object_text):
        pass
        #self.label.setText(f'{object_name},  {object_text}')
        #print(f'{object_name}  {object_text}')

    def find_button_reaction(self):
        self.results_counter = 0
        self.filter_values=[]
        for base_name in self.values:
            self.filter_values.append([base_name, self.findChild(QComboBox, base_name).currentText()])

        self.coincidence_quantity = 0 
        self.coincidences_out = ''
        self.exit_switch = 0
        self.line_draw_coordinates = []
        self.line_draw_dict = {}
        self.line_drawed = False

        for row_number in range(2, working_sheet.max_row+1):
            self.coincidence_quantity = 0
            

            for column_number in range(1, len(self.values)+1):

                self.filter_value = self.filter_values[column_number-1][1]
                self.sheet_value = working_sheet.cell(row=row_number, column=column_number).value 

                if isinstance(self.sheet_value, int) or isinstance(self.sheet_value, float):
                    self.sheet_value = str(round(self.sheet_value, 2))

                if self.filter_value == self.sheet_value:
                    self.coincidence_quantity+=1

                elif self.filter_value == None or self.filter_value == working_sheet.cell(row=1, column=column_number).value:
                    self.coincidence_quantity+=1

                else:
                    break

            self.appender=''

            if self.coincidence_quantity == len(self.values):
                #print(self.coincidence_quantity, self.values, len(self.values))
                self.results_counter+=1
                self.appender = f'\n----------{self.results_counter}----------\n'

                # КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ
                if self.filter_values[2][1] not in self.values:      # if self.filter_values[7][1] not in self.values
                        
                    if self.filter_values[7][1] not in self.values:
                        self.start_cords = None
                        self.end_cords = None

                        for cords in self.coordinates:
                            if self.filter_values[2][1] in cords:
                                self.start_cords = (cords[0], cords[1])

                            elif self.filter_values[7][1] in cords:
                                self.end_cords = (cords[0], cords[1])
                            
                        if self.start_cords != None and self.end_cords != None:
                            #self.line_points = []
                            #self.line_points.append(self.start_cords)
                            #self.line_points.append(self.end_cords)

                            #print(self.line_points)

                            #PolyLine(self.line_points, color="red", weight=2.5, opacity=3).add_to(self.map_)
                            #self.line_draw_coordinates.append(self.line_points)
                            self.line_draw_dict.update({0 : [self.start_cords, self.end_cords]})
                            self.line_drawed = True
                            print('drawed with all coordinates')
                            
                        else:
                            msg_showing('Ошибка', f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            
                            print(f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            break

                else:
                    if self.filter_values[7][1] not in self.values:
                        #print('We have end coordinate')
                        
                        self.find_start_cords=working_sheet.cell(row=row_number, column=3).value
                        self.find_end_cords = self.filter_values[7][1]
                        self.start_cords = None
                        self.end_cords = None

                        for cords in self.coordinates:
                            if self.find_start_cords == cords[2]:
                                self.start_cords = (cords[0], cords[1])

                            elif self.find_end_cords == cords[2]:
                                self.end_cords = (cords[0], cords[1])
                        
                        if self.start_cords != None and self.end_cords != None:
                            
                            self.line_draw_dict.update({0 : [self.start_cords, self.end_cords]})
                            #PolyLine(self.line_points, color="red", weight=2.5, opacity=3).add_to(self.map_)
                            #print('Line drawed')
                            self.line_drawed = True

                        else:
                            msg_showing('Ошибка', f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            
                            print(f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            break

                    else:
                        msg_showing('Предупреждение', 'Не было выбрано ни одной страны для отрисовки указателей')
                        self.exit_switch = 1
                        break

                # КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ

                for column_number in range(1, len(self.values)):
                    self.sheet_value = working_sheet.cell(row=row_number, column=column_number).value

                    if isinstance(self.sheet_value, int) or isinstance(self.sheet_value, float):
                        self.sheet_value = str(round(self.sheet_value, 2))

                    self.appender+=working_sheet.cell(row=1, column=column_number).value + ' : ' + self.sheet_value + '\n'
                self.appender+='------------------------------\n'


            self.coincidences_out+=self.appender


            if self.exit_switch:
                break

        data = io.BytesIO()                 # Не спрашивайте, так было на стаке
        self.map_window.close()
        
        #map1.save(data, close_file=False)   # Тут тоже

        self.map_window = MapWindow(data)
        #self.map_window.setWindowTitle('Карта Мира')

        #print(self.line_drawed)

        if self.line_drawed == True:
            print('Making marker')
            self.depart_coords = self.line_draw_dict[0][0]
            PolyLine(self.line_draw_dict[0], color="red", weight=2.5, opacity=3).add_to(self.map_)
            map_update(self.depart_coords[0], self.depart_coords[1], self.coincidences_out, self.map_, 'purple')

        self.map_window.show()
        
        #self.widget.setParent(None)
        self.map_.save(data, close_file=False)

        self.scrollAreaWidgetContents = QLabel(f'Найдено совпадений: {self.results_counter-1}' + '\n' + str(self.coincidences_out))   # Изменяем текст прокручиваемого окна на найденные страны
        self.area.setWidget(self.scrollAreaWidgetContents)


        
        #print(f'\nfilter values :\n{self.filter_values}\nnumber of concidences : {len(self.coincidences_out)}\nconcidences list :\n{self.coincidences_out}\n')

    def window_construct(self):         # Для трех строчек надо сильно думать, ограничимся двумя. Values - спсок названий стобцов.
        self.switch_column=0


        if self.rows_count >= 3:
            msg_showing('Ошибка', 'Программа не поддерживает разбиение кнопок на три и более строки.')                  # Нюансек
            sys.exit(0)

        else:
            self.grid_layout=QGridLayout(self)

            self.move_column = len(self.values) - len(self.values) // self.rows_count
            row_move=0

            for column_number, obj_name in enumerate(self.values, start = 0):           # Перебор всех значений на входе. В данном случае - названий нужныз столбцов
                self.values_list=[]
                self.combo = QComboBox(self)
                self.combo.setObjectName(str(obj_name))
                

                for row_number in range(2, working_sheet.max_row+1):
                    if row_number == 2:
                        self.combo.addItem(obj_name)

                    else:
                        check=working_sheet.cell(row=row_number, column=column_number+1).value

                        if isinstance(check, int) or isinstance(check, float):
                            check=str(round(check, 2))

                        if check not in self.values_list:
                            self.values_list.append(check)

                self.values_list.sort()

                for obj in self.values_list:
                    self.combo.addItem(obj)

                if self.move_column == column_number:
                    row_move+=1
                    self.grid_layout.addWidget(self.combo, row_move, 0)
                    self.switch_column=1
                    
                else:
                    self.grid_layout.addWidget(self.combo, row_move, self.switch_column)
                    self.switch_column+=1

                self.combo.currentIndexChanged.connect(lambda ch, name=self.combo.objectName() : self.index_change_reaction(name, self.findChild(QComboBox, name).currentText()))
                #self.combo.currentIndexChanged.connect(lambda ch, name=self.combo.objectName() : self.index_change_reaction(name, self.combo.currentText()))
        self.find_button = QPushButton(self)
        self.find_button.setText('Поиск')
        self.find_button.clicked.connect(self.find_button_reaction)

        self.area = QScrollArea(self)               # Создание объекта, способного реализовывать прокрутку своего содержимого. При множестве найденных результатов поиска это - лучшее решение
        #self.area.setFont(font)                     # Форматируем объект. В данном случае - только меняем размер шрифта
        self.area.setWidgetResizable(True)          # Говорим проге, что содержимое можно прокручивать

        self.grid_layout.addWidget(self.find_button, 3, 0)  
        self.grid_layout.addWidget(self.area, 3, 1, 2, self.move_column)

class MapWindow(QWidget):
    def __init__(self, buffer_data, is_refresh=0):                                # "Магический" метод, позволяющий выполнять указанные действия при создании экземпляров класса
        super().__init__()
        self.buffer_data = buffer_data
        self.screen_w = screen_w
        self.screen_h = screen_h
        self.resize(screen_w, screen_h)         # Ресайзим окно по ШиринеХВысоте монитора. 
        font = QtGui.QFont()                            # Указываем шрифт текста
        font.setPointSize(14)

        if is_refresh:
            self.refresh_map()
        else:
            self.initUI()                               # При создании экземпляра вызываем функцию отрисовки интерфейса
        

    def refresh_map(self):
        print(map_window.findChild(QWebEngineView, 'map_view').objectName())
        print(map_window.findChild(QGridLayout, 'map_grid_layout').objectName())
        

        #map_window.grid_layout = QGridLayout(map_window)
        #map_window.grid_layout.setObjectName('map_grid_layout')

        #map_window.map_view = QWebEngineView(map_window)            # Создание объекта для просмотра html-файла нашей карты
        #map_window.map_view.setObjectName('map_view')

        map_window.findChild(QWebEngineView, 'map_view').setParent(None)
        map_window.findChild(QGridLayout, 'map_grid_layout').setParent(None)

        map_window.map_view.setHtml(map_window.buffer_data.getvalue().decode())
        map_window.map_view.setWindowTitle('Карта Мира')
        map_window.grid_layout.addWidget(map_window.map_view, 5, 0, 10, (len(values_dict) - len(values_dict) // len(values_dict)))



    def initUI(self):                               # Функция для отрисовки интерфейса
        self.grid_layout = QGridLayout(self)
        self.grid_layout.setObjectName('map_grid_layout')

        self.map_view = QWebEngineView(self)            # Создание объекта для просмотра html-файла нашей карты
        self.map_view.setObjectName('map_view')

        self.map_view.setHtml(self.buffer_data.getvalue().decode())
        self.map_view.setWindowTitle('Карта Мира')
        self.grid_layout.addWidget(self.map_view, 5, 0, 10, (len(values_dict) - len(values_dict) // len(values_dict)))

        '''print(self.findChild(QWebEngineView, 'map_view').objectName())
        print(self.findChild(QGridLayout, 'map_grid_layout').objectName())'''


for row_number in range(3, 241+1):  # Проходимся по столбцам справа таблицы
    countries.append([working_sheet.cell(row=row_number, column=28).value, working_sheet.cell(row=row_number, column=29).value, working_sheet.cell(row=row_number, column=30).value])
countries.sort()    # Сортируем страны в алфавитном порядке


for list_ in countries:     # Перебираем списки из массива, созданного выше
    value = list_[2]
    if value!=None:
        first_cord=float(value[ : value.index(',')])
        second_cord=float(value[value.index(',')+2 : ])
        
        coordinates.append([first_cord, second_cord, list_[0]])
        

    

map1=Map(location=[0, 0], zoom_start=3)     # Создаем объект карты (поскольку мы прописали from folium import *, то есть, импортировали все содержимое главного модуля folium, а метод Map относится именно к нему, то мы можем не прописывать folium.Map, а писать сразу Map)

#for country_data in coordinates:            # Поочередно выбираем списки с информацией для построения маркеров
    #map_update(country_data[0], country_data[1], country_data[2], map1, 'red')      # Вызываем функцию доя нанесения маркера на карту, передавая ей координаты (2), название страны, объект карты и цвет маркера
#map1.save('map.html', close_file=False) # Сохранение карты в директории, где находится питоновский файл



app = QApplication(sys.argv)

column_dicts()


screen = app.primaryScreen().availableGeometry()        # Получаем значения доступного для использования пространства монитора 
screen_w=screen.width()                                 # Записываем доступную ширину монитора
screen_h=screen.height()                                # Записываем доступную высоту монитора

data = io.BytesIO()             # Не спрашивайте, так было на стаке

save_path = list(__file__)
save_path.reverse()
save_path = save_path[save_path.index('\\') : ]
save_path.reverse()
save_path_add = ''
for symbol in save_path:
    save_path_add+=symbol

save_path = save_path_add+'map.html'

map1.save(data, close_file=False)   # Тут тоже

map_window = MapWindow(data)
map_window.setObjectName('MapWindow')
map_window.setWindowTitle('Карта Мира')
map_window.show()

#print(values_dict)
#print(coordinates)
main_window = mainWindow(values_dict, coordinates, map1, map_window)                   # Создаем экземпляр класса ResultWindow, отвечающего за выведение окна для поиска и выведения результатов поиска
main_window.setObjectName('mainWindow')             # Присваиваем экземпляру внутреннее программное имя
main_window.setWindowTitle('Поиск')                     # Меняем заголовок окна
main_window.show()                                      # Выводим диалоговое окно

sys.exit(app.exec_())










import openpyxl     # pip install openpyxl , библа для работы с таблицей
from openpyxl import *      # вытаскиваем из нее все вложенные методы
import folium   # pip install folium , библа для работы с картой
from folium import *    # вытаскиваем из нее все вложенные методы
import io   # Библиотека для отрисовки карты в проге
import os
from os import *
import sys
from PyQt5 import QtGui
from PyQt5.QtWidgets import *   # pip install pyqt5 , библиотека для создания интерфейса
from PyQt5.QtWebEngineWidgets import *      # pip install PyQtWebEngine
from PyQt5.QtCore import *

values_dict={}
countries=[]
coordinates=[]
mid_list=[]

_translate=QCoreApplication.translate


vadim_s_build=True      # !!!!!!!! ЕСЛИ У ВАС В ПРОГЕ ЭТА ПЕРЕМЕННАЯ True, ПОМЕНЯЙТЕ ЕЕ НА False ИЛИ ЗАМЕНИТЕ ПУТЬ ПРИ ВЫПОЛНЕНИИ УСЛОВИЯ НА СВОЙ !!!!!!!!

if vadim_s_build:
    xl_path='J:/Downloads/Telegram Desktop/Транзит 2019-2020 гг..xlsx'
else:
    xl_path=input('Введите полный путь до Excel-таблицы (вместе с расширением) ')

xl_file = load_workbook(filename=xl_path,  data_only=True)  # Создание мелкой копии файла. Атрибут data_only позволяет избежать получения формул при попытке получить данные из ячейки
working_sheet=xl_file[xl_file.sheetnames[0]]    # Выбор таблицы. xl_file.sheetnames возвращает список с листами таблицы. Указывая [0] мы передаем проге инфу название нужного листа и он полнстью записывается в working_sheet
max_rows=working_sheet.max_row

def msg_showing(main_text, secondary_text):     # Функция для вызова информационного окна с указанными заголовком и основным текстом
    msg = QMessageBox()
    msg.setWindowTitle(main_text)
    msg.setText(secondary_text)
    #msg.setIcon(QMessageBox.Warning)
    msg.exec_()

def sizeHint(widget):
    a = widget.sizeHint()
    a.setHeight(250)
    a.setWidth(50)
    return a

def map_update(x, y, popup, map_, color, zoom=3):   # Функция для добавления маркеров по указанным координатам на карту, так же указанную при вызове функции
    return(folium.Marker(location=[x, y], icon=folium.Icon(color = color.lower()), popup=popup).add_to(map_))   
    #return(Map(location=[x, y], zoom_start=zoom))

def column_dicts():

    for column_number in range(1, working_sheet.max_column+1):
        if working_sheet.cell(row=1, column=column_number).value == None:
            break
        else:
            column_values=[]
            for row_number in range(2, working_sheet.max_row+1):
                value=working_sheet.cell(row=row_number, column=column_number).value

                if value == None:
                    break

                else:
                    if value not in column_values:
                        if isinstance(value, float):
                            value=round(value, 2)
                        column_values.append(value)
            column_values.sort()
            values_dict.update({working_sheet.cell(row=1, column=column_number).value : column_values})

class mainWindow(QWidget):
    def __init__(self, values, coordinates, map_, map_window): # Названия столбоцв
        super().__init__()
        self.rows_count=2
        self.values = values 
        self.coordinates = coordinates
        self.map_ = map_
        self.map_window = map_window
        self.resize(int(screen_w/2), int(screen_h/2))  # Ресайз окна под половину ширины и высоты монитора
        #self.resize(int(1920/2), int(1080/2))

        self.window_construct()

    def index_change_reaction(self, object_name, object_text):
        pass
        #self.label.setText(f'{object_name},  {object_text}')
        #print(f'{object_name}  {object_text}')

    def find_button_reaction(self):
        self.results_counter = 0
        self.filter_values=[]
        for base_name in self.values:
            self.filter_values.append([base_name, self.findChild(QComboBox, base_name).currentText()])

        self.coincidence_quantity = 0 
        self.coincidences_out = ''
        self.exit_switch = 0
        self.line_draw_coordinates = []

        for row_number in range(2, working_sheet.max_row+1):
            self.coincidence_quantity = 0
            self.line_drawed = False

            for column_number in range(1, len(self.values)+1):

                self.filter_value = self.filter_values[column_number-1][1]
                self.sheet_value = working_sheet.cell(row=row_number, column=column_number).value 

                if isinstance(self.sheet_value, int) or isinstance(self.sheet_value, float):
                    self.sheet_value = str(round(self.sheet_value, 2))

                if self.filter_value == self.sheet_value:
                    self.coincidence_quantity+=1

                elif self.filter_value == None or self.filter_value == working_sheet.cell(row=1, column=column_number).value:
                    self.coincidence_quantity+=1

                else:
                    break

            self.appender=''

            if self.coincidence_quantity == len(self.values):
                
                #print(self.coincidence_quantity, self.values, len(self.values))
                self.results_counter+=1
                self.appender = f'\n----------{self.results_counter}----------\n'

                # КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ
                if self.filter_values[2][1] not in self.values:      # if self.filter_values[7][1] not in self.values
                        
                    if self.filter_values[7][1] not in self.values:
                        self.start_cords = None
                        self.end_cords = None

                        for cords in self.coordinates:
                            if self.filter_values[2][1] in cords:
                                self.start_cords = (cords[0], cords[1])

                            elif self.filter_values[7][1] in cords:
                                self.end_cords = (cords[0], cords[1])
                            
                        if self.start_cords != None and self.end_cords != None:
                            self.line_points = []
                            self.line_points.append(self.start_cords)
                            self.line_points.append(self.end_cords)

                            #print(self.line_points)

                            #PolyLine(self.line_points, color="red", weight=2.5, opacity=3).add_to(self.map_)
                            self.line_draw_coordinates.append(self.line_points)
                            self.line_drawed = True
                            
                        else:
                            msg_showing('Ошибка', f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            
                            print(f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            break

                else:
                    if self.filter_values[7][1] not in self.values:
                        #print('We have end coordinate')
                        
                        self.find_start_cords=working_sheet.cell(row=row_number, column=3).value
                        self.find_end_cords = self.filter_values[7][1]
                        self.start_cords = None
                        self.end_cords = None

                        for cords in self.coordinates:
                            if self.find_start_cords == cords[2]:
                                self.start_cords = (cords[0], cords[1])

                            elif self.find_end_cords == cords[2]:
                                self.end_cords = (cords[0], cords[1])
                        
                        if self.start_cords != None and self.end_cords != None:
                            self.line_points = []
                            self.line_points.append(self.start_cords)
                            self.line_points.append(self.end_cords)

                            self.line_draw_coordinates.append(self.line_points)

                            #PolyLine(self.line_points, color="red", weight=2.5, opacity=3).add_to(self.map_)
                            #print('Line drawed')
                            self.line_drawed = True

                        else:
                            msg_showing('Ошибка', f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            
                            print(f"В таблице не оказалось координат для построения указателя для пути \" {self.filter_values[2][1]} ---> {self.filter_values[7][1]} \"")
                            break

                    else:
                        msg_showing('Предупреждение', 'Не было выбрано ни одной страны для отрисовки указателей')
                        self.exit_switch = 1
                        break

                # КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ КОСТЫЛЬ

                for column_number in range(1, len(self.values)):
                    self.sheet_value = working_sheet.cell(row=row_number, column=column_number).value

                    if isinstance(self.sheet_value, int) or isinstance(self.sheet_value, float):
                        self.sheet_value = str(round(self.sheet_value, 2))

                    self.appender+=working_sheet.cell(row=1, column=column_number).value + ' : ' + self.sheet_value + '\n'
                self.appender+='------------------------------\n'


            self.coincidences_out+=self.appender


            if self.exit_switch:
                break

            
        self.map_window.close()
        data = io.BytesIO()                 # Не спрашивайте, так было на стаке
        map1.save(data, close_file=False)   # Тут тоже

        self.map_window = MapWindow(data)
        self.map_window.setWindowTitle('Карта Мира')

        if self.line_drawed:
            print('Making marker')
            self.depart_coords = self.line_draw_coordinates[0][0]
                    
            #self.coord_begin = self.line_points[0]
            #print(self.coord_begin)
            #self.coord_end = self.line_points[0][1]
            map_update(self.depart_coords[0], self.depart_coords[1], self.coincidences_out, map1, 'purple')

        self.map_window.show()

        #self.widget.setParent(None)
        self.map_.save('map.html', close_file=False)

        self.scrollAreaWidgetContents = QLabel(f'Найдено совпадений: {self.results_counter-1}' + '\n' + str(self.coincidences_out))   # Изменяем текст прокручиваемого окна на найденные страны
        self.area.setWidget(self.scrollAreaWidgetContents)


        
                #print(f'\nfilter values :\n{self.filter_values}\nnumber of concidences : {len(self.coincidences_out)}\nconcidences list :\n{self.coincidences_out}\n')

    def window_construct(self):         # Для трех строчек надо сильно думать, ограничимся двумя. Values - спсок названий стобцов.
        self.switch_column=0


        if self.rows_count >= 3:
            msg_showing('Ошибка', 'Программа не поддерживает разбиение кнопок на три и более строки.')                  # Нюансек
            sys.exit(0)

        else:
            self.grid_layout=QGridLayout(self)

            self.move_column = len(self.values) - len(self.values) // self.rows_count
            row_move=0

            for column_number, obj_name in enumerate(self.values, start = 0):           # Перебор всех значений на входе. В данном случае - названий нужныз столбцов
                self.values_list=[]
                self.combo = QComboBox(self)
                self.combo.setObjectName(str(obj_name))
                

                for row_number in range(2, working_sheet.max_row+1):
                    if row_number == 2:
                        self.combo.addItem(obj_name)

                    else:
                        check=working_sheet.cell(row=row_number, column=column_number+1).value

                        if isinstance(check, int) or isinstance(check, float):
                            check=str(round(check, 2))

                        if check not in self.values_list:
                            self.values_list.append(check)

                self.values_list.sort()

                for obj in self.values_list:
                    self.combo.addItem(obj)

                if self.move_column == column_number:
                    row_move+=1
                    self.grid_layout.addWidget(self.combo, row_move, 0)
                    self.switch_column=1
                    
                else:
                    self.grid_layout.addWidget(self.combo, row_move, self.switch_column)
                    self.switch_column+=1

                self.combo.currentIndexChanged.connect(lambda ch, name=self.combo.objectName() : self.index_change_reaction(name, self.findChild(QComboBox, name).currentText()))
                #self.combo.currentIndexChanged.connect(lambda ch, name=self.combo.objectName() : self.index_change_reaction(name, self.combo.currentText()))
        self.find_button = QPushButton(self)
        self.find_button.setText('Поиск')
        self.find_button.clicked.connect(self.find_button_reaction)

        self.area = QScrollArea(self)               # Создание объекта, способного реализовывать прокрутку своего содержимого. При множестве найденных результатов поиска это - лучшее решение
        #self.area.setFont(font)                     # Форматируем объект. В данном случае - только меняем размер шрифта
        self.area.setWidgetResizable(True)          # Говорим проге, что содержимое можно прокручивать

        self.grid_layout.addWidget(self.find_button, 3, 0)  
        self.grid_layout.addWidget(self.area, 3, 1, 2, self.move_column)

class MapWindow(QWidget):
    def __init__(self, buffer_data, is_refresh=0):                                # "Магический" метод, позволяющий выполнять указанные действия при создании экземпляров класса
        super().__init__()
        self.buffer_data = buffer_data
        self.screen_w = screen_w
        self.screen_h = screen_h
        self.resize(screen_w, screen_h)         # Ресайзим окно по ШиринеХВысоте монитора. 
        font = QtGui.QFont()                            # Указываем шрифт текста
        font.setPointSize(14)

        if is_refresh:
            self.refresh_map()
        else:
            self.initUI()                               # При создании экземпляра вызываем функцию отрисовки интерфейса
        

    def refresh_map(self):
        print(map_window.findChild(QWebEngineView, 'map_view').objectName())
        print(map_window.findChild(QGridLayout, 'map_grid_layout').objectName())
        

        #map_window.grid_layout = QGridLayout(map_window)
        #map_window.grid_layout.setObjectName('map_grid_layout')

        #map_window.map_view = QWebEngineView(map_window)            # Создание объекта для просмотра html-файла нашей карты
        #map_window.map_view.setObjectName('map_view')

        map_window.findChild(QWebEngineView, 'map_view').setParent(None)
        map_window.findChild(QGridLayout, 'map_grid_layout').setParent(None)

        map_window.map_view.setHtml(map_window.buffer_data.getvalue().decode())
        map_window.map_view.setWindowTitle('Карта Мира')
        map_window.grid_layout.addWidget(map_window.map_view, 5, 0, 10, (len(values_dict) - len(values_dict) // len(values_dict)))



    def initUI(self):                               # Функция для отрисовки интерфейса
        self.grid_layout = QGridLayout(self)
        self.grid_layout.setObjectName('map_grid_layout')

        self.map_view = QWebEngineView(self)            # Создание объекта для просмотра html-файла нашей карты
        self.map_view.setObjectName('map_view')

        self.map_view.setHtml(self.buffer_data.getvalue().decode())
        self.map_view.setWindowTitle('Карта Мира')
        self.grid_layout.addWidget(self.map_view, 5, 0, 10, (len(values_dict) - len(values_dict) // len(values_dict)))

        '''print(self.findChild(QWebEngineView, 'map_view').objectName())
        print(self.findChild(QGridLayout, 'map_grid_layout').objectName())'''


for row_number in range(3, 241+1):  # Проходимся по столбцам справа таблицы
    countries.append([working_sheet.cell(row=row_number, column=28).value, working_sheet.cell(row=row_number, column=29).value, working_sheet.cell(row=row_number, column=30).value])
countries.sort()    # Сортируем страны в алфавитном порядке


for list_ in countries:     # Перебираем списки из массива, созданного выше
    value = list_[2]
    if value!=None:
        first_cord=float(value[ : value.index(',')])
        second_cord=float(value[value.index(',')+2 : ])
        
        coordinates.append([first_cord, second_cord, list_[0]])
        

    

map1=Map(location=[0, 0], zoom_start=3)     # Создаем объект карты (поскольку мы прописали from folium import *, то есть, импортировали все содержимое главного модуля folium, а метод Map относится именно к нему, то мы можем не прописывать folium.Map, а писать сразу Map)

#for country_data in coordinates:            # Поочередно выбираем списки с информацией для построения маркеров
    #map_update(country_data[0], country_data[1], country_data[2], map1, 'red')      # Вызываем функцию доя нанесения маркера на карту, передавая ей координаты (2), название страны, объект карты и цвет маркера
#map1.save('map.html', close_file=False) # Сохранение карты в директории, где находится питоновский файл



app = QApplication(sys.argv)

column_dicts()


screen = app.primaryScreen().availableGeometry()        # Получаем значения доступного для использования пространства монитора 
screen_w=screen.width()                                 # Записываем доступную ширину монитора
screen_h=screen.height()                                # Записываем доступную высоту монитора

data = io.BytesIO()             # Не спрашивайте, так было на стаке

save_path = list(__file__)
save_path.reverse()
save_path = save_path[save_path.index('\\') : ]
save_path.reverse()
save_path_add = ''
for symbol in save_path:
    save_path_add+=symbol

save_path = save_path_add+'map.html'

map1.save(data, close_file=False)   # Тут тоже

map_window = MapWindow(data)
map_window.setObjectName('MapWindow')
map_window.setWindowTitle('Карта Мира')
map_window.show()

#print(values_dict)
#print(coordinates)
main_window = mainWindow(values_dict, coordinates, map1, map_window)                   # Создаем экземпляр класса ResultWindow, отвечающего за выведение окна для поиска и выведения результатов поиска
main_window.setObjectName('mainWindow')             # Присваиваем экземпляру внутреннее программное имя
main_window.setWindowTitle('Поиск')                     # Меняем заголовок окна
main_window.show()                                      # Выводим диалоговое окно

sys.exit(app.exec_())